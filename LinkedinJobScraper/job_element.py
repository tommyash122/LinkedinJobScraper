import openpyxl
import os
from datetime import date

class JobElement:
    def __init__(self, job_description, company, location, link):
        self.job_description = job_description
        self.company = company
        self.location = location
        self.date_submitted = date.today().strftime('%x')
        self.link = link
        
    def getAsDictionary(self):
        return {
            "job_description": self.job_description,
            "company": self.company,
            "location": self.location,
            "date_submitted": self.date_submitted,
            "link": self.link      
        }

def update_excel(objects_list):
    # Create backup of current Excel file
    backup_path = "jobs_backup.xlsx"
    if os.path.exists("jobs.xlsx"):
        os.replace("jobs.xlsx", backup_path)

    # Open or create Excel file and sheet
    try:
        wb = openpyxl.load_workbook("jobs.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Job Listings"
    
    # Write header row if necessary
    if not sheet["A1"].value:
        sheet["A1"] = "Job Description"
        sheet["B1"] = "Company"
        sheet["C1"] = "Location"
        sheet["D1"] = "Date Submitted"
        sheet["E1"] = "Link"

    # Write data rows
    for i, obj in enumerate(objects_list, start=2):
        row_num = str(i)
        sheet.cell(row=i, column=1).value = obj.job_description
        sheet.cell(row=i, column=2).value = obj.company
        sheet.cell(row=i, column=3).value = obj.location
        sheet.cell(row=i, column=4).value = obj.date_submitted
        sheet.cell(row=i, column=5).value = obj.link

    # Save Excel file
    try:
        wb.save("jobs.xlsx")
        print(f"Successfully updated {len(objects_list)} job listings.")
    except Exception as e:
        print(f"An error occurred while saving the Excel file: {e}")
        # Restore backup
        if os.path.exists(backup_path):
            os.replace(backup_path, "jobs.xlsx")
            print("Excel file restored from backup.")
        raise
